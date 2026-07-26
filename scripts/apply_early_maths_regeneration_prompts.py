#!/usr/bin/env python3
"""Apply crop-safe regeneration prompts to the Early Maths LKG master workbook.

The script updates only the pages listed in the committed regeneration prompt pack.
It preserves all other workbook rows and writes a new workbook by default.
"""
from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PACK = ROOT / "production-prompts" / "early-maths-adventures" / "lkg" / "v4" / "regeneration" / "early-maths-crop-safe-regeneration-prompts-v1.json"


def load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"JSON object expected: {path}")
    return value


def build_prompt(page_id: str, page: dict[str, Any], global_rules: list[str], book: str, level: str) -> str:
    assets = "\n".join(f"- {name}: {description}" for name, description in page["named_assets"].items())
    rules = "\n".join(f"- {rule}" for rule in global_rules)
    crop_names = ", ".join(page["named_assets"])
    return f"""BCube Crop-Safe Regeneration Illustration Prompt — {page_id}

BOOK AND LEVEL
Book: {book}
Level: {level}
Exact page title: {page['title']}
Regeneration purpose: replace the previous source sheet with a deterministic crop-safe illustration-only asset sheet.
Asset layout: {page['layout']}

ILLUSTRATION ROLE — LOCKED
Create exactly one illustration-only source sheet containing exactly the named assets below. The publishing engine will crop these assets and add every worksheet mechanic, including all text, numerals not explicitly named, response controls, lines, arrows, boxes, panels, choices, teacher cue, branding and page number.

EXACT NAMED ASSETS
{assets}

PAGE-SPECIFIC CROP LOCK
{page['page_specific_lock']}

GLOBAL CROP-SAFE LOCK
{rules}

EXACT CROP NAMES
Use these names without renaming, merging, duplicating or omitting any asset: {crop_names}.

ACCEPTANCE GATE
Reject the illustration if any asset is clipped, too close to a neighbour, merged with another asset, surrounded by worksheet mechanics, semantically different from its description, or not independently crop-safe. Leave noticeably more white space than visually necessary. No alternate, mockup, contact sheet, explanation or second image."""


def main() -> int:
    parser = argparse.ArgumentParser(description="Patch Early Maths LKG regeneration prompts into the master workbook")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--sheet", default="All Page Prompts")
    parser.add_argument("--in-place", action="store_true")
    args = parser.parse_args()

    source = args.workbook.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)

    if args.in_place:
        output = source
        backup = source.with_suffix(source.suffix + ".before-regeneration-prompts.bak")
        shutil.copy2(source, backup)
    else:
        output = (args.output or source.with_name(source.stem + "_Early_Maths_Regeneration_Prompts.xlsx")).expanduser().resolve()

    pack = load_json(PACK)
    pages = pack.get("pages", {})
    if not pages:
        raise ValueError("Regeneration prompt pack contains no pages")

    wb = load_workbook(source)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]
    headers = {str(cell.value).strip(): i + 1 for i, cell in enumerate(ws[1]) if cell.value is not None}
    required = ["Prompt ID", "Complete Standalone Illustration Prompt", "Output Filename", "Status", "Ready to Test", "Prompt Validation Status"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing workbook columns: {missing}")

    found: set[str] = set()
    for row in range(2, ws.max_row + 1):
        page_id = str(ws.cell(row, headers["Prompt ID"]).value or "").strip()
        page = pages.get(page_id)
        if page is None:
            continue
        prompt = build_prompt(page_id, page, pack["global_crop_safe_lock"], pack["book"], pack["level"])
        ws.cell(row, headers["Complete Standalone Illustration Prompt"]).value = prompt
        ws.cell(row, headers["Output Filename"]).value = f"{page_id}.png"
        ws.cell(row, headers["Status"]).value = "REGENERATION PROMPT UPDATED — CROP SAFE"
        ws.cell(row, headers["Ready to Test"]).value = "REGENERATE ILLUSTRATION"
        ws.cell(row, headers["Prompt Validation Status"]).value = "CROP-SAFE REGENERATION REQUIRED"
        if "Asset Layout Standard" in headers:
            ws.cell(row, headers["Asset Layout Standard"]).value = page["layout"]
        if "Crop-Safe Spacing Rule" in headers:
            ws.cell(row, headers["Crop-Safe Spacing Rule"]).value = "Minimum 10% inter-zone white gutter and 8% clear padding around each named asset; no worksheet mechanics or neighbouring-object contamination."
        found.add(page_id)

    missing_pages = sorted(set(pages) - found)
    if missing_pages:
        raise ValueError(f"Workbook does not contain regeneration pages: {missing_pages}")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(json.dumps({
        "output": str(output),
        "updated_pages": len(found),
        "page_ids": sorted(found),
        "source_pack": str(PACK),
        "policy": "Only listed regeneration pages were changed; all other workbook rows were preserved."
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
