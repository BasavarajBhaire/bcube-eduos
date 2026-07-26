#!/usr/bin/env python3
"""Rebuild every Early Maths Adventures LKG illustration prompt and page blueprint.

The workbook remains the single source of truth. For every page this script writes
one illustration-only asset specification, one deterministic page-execution
contract, one named crop manifest and one matching layout standard. Asset names
are repeated in both prompts so the illustration and runtime contract cannot
drift silently.
"""
from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BOOK = "Early Maths Adventures"
LEVEL = "LKG"
SLUG = "early-maths-adventures"


def crops_grid(names: list[str], cols: int) -> dict[str, list[float]]:
    rows = (len(names) + cols - 1) // cols
    gap_x, gap_y = 0.04, 0.05
    usable_w = 1 - gap_x * (cols + 1)
    usable_h = 1 - gap_y * (rows + 1)
    cw, ch = usable_w / cols, usable_h / rows
    out: dict[str, list[float]] = {}
    for i, name in enumerate(names):
        r, c = divmod(i, cols)
        x0 = gap_x + c * (cw + gap_x)
        y0 = gap_y + r * (ch + gap_y)
        out[name] = [round(x0, 3), round(y0, 3), round(x0 + cw, 3), round(y0 + ch, 3)]
    return out


def spec(title: str, objective: str, instruction: str, archetype: str, mechanic: str,
         layout: str, assets: dict[str, str], teacher: str, *, cols: int = 3,
         response: str = "Use only the page-specific response controls defined below.") -> dict[str, Any]:
    return {
        "title": title,
        "objective": objective,
        "instruction": instruction,
        "archetype": archetype,
        "mechanic": mechanic,
        "layout": layout,
        "assets": assets,
        "crops": crops_grid(list(assets), cols),
        "teacher": teacher,
        "response": response,
    }


SPECS: dict[int, dict[str, Any]] = {
    8: spec("Numbers 1–10 Review", "Recognise and use numbers 1–10.",
            "Count each set and match or circle the correct numeral.", "A07 Count & Record",
            "count-match-review-1-10", "count-review-ten-cards-v1",
            {"set_1":"one red apple", "set_2":"two blue birds", "set_3":"three yellow stars", "set_4":"four green leaves", "set_5":"five orange fish", "set_6":"six purple flowers", "set_7":"seven toy blocks", "set_8":"eight balloons", "set_9":"nine ladybirds", "set_10":"ten pencils"},
            "Ask the child to touch each object once while counting, then choose the numeral.", cols=5),
    9: spec("Numbers 11–20", "Recognise numbers 11–20.",
            "Count each organised set and choose the matching numeral from 11 to 20.", "A07 Count & Record",
            "count-select-11-20", "count-11-20-five-cards-v1",
            {"set_11":"eleven beads in two clear rows", "set_13":"thirteen shells in two clear rows", "set_15":"fifteen counters in three rows", "set_18":"eighteen stars in three rows", "set_20":"twenty blocks in four rows"},
            "Use row counting and pause after ten before continuing.", cols=5),
    10: spec("Count & Match", "Count accurately and match quantity to numeral.",
             "Count each object group. Draw a line to its matching numeral.", "A05 Read/Look & Match",
             "count-and-match", "matching-assets-grid-v2",
             {"left_1":"group of two kites", "left_2":"group of four ducks", "left_3":"group of six cupcakes", "left_4":"group of eight balls", "right_1":"large numeral 6 token without text styling", "right_2":"large numeral 2 token without card", "right_3":"large numeral 8 token without card", "right_4":"large numeral 4 token without card"},
             "Model counting the first group, then let the child match the remaining groups.", cols=2),
    11: spec("Count & Circle", "Identify the correct quantity.",
             "Count each group. Circle the correct numeral below it.", "A07 Count & Record",
             "count-and-circle-number", "count-circle-six-groups-v1",
             {"apple_group":"one apple", "star_group":"two stars", "balloon_group":"three balloons", "fish_group":"four fish", "flower_group":"five flowers", "leaf_group":"six leaves"},
             "Touch each object once while counting aloud, then circle the matching numeral.", cols=3),
    12: spec("More or Less", "Compare quantities.",
             "Look at each pair. Circle the group with more or less as instructed.", "A08 Compare & Choose",
             "compare-more-less", "comparison-three-pairs-v1",
             {"pair_1_left":"two red apples", "pair_1_right":"five red apples", "pair_2_left":"six blue fish", "pair_2_right":"three blue fish", "pair_3_left":"four yellow stars", "pair_3_right":"seven yellow stars"},
             "Ask the child to compare without recounting first, then verify by counting.", cols=2),
    13: spec("Equal Groups", "Recognise equal groups.",
             "Count both groups in each pair. Circle the pairs that are equal.", "A08 Compare & Choose",
             "identify-equal-groups", "equal-groups-three-pairs-v1",
             {"pair_1_left":"four oranges", "pair_1_right":"four oranges", "pair_2_left":"three butterflies", "pair_2_right":"five butterflies", "pair_3_left":"six blocks", "pair_3_right":"six blocks"},
             "Count one group, then check whether the partner group has the same number.", cols=2),
    14: spec("Missing Numbers", "Complete number sequences.",
             "Look at each number sequence. Write or choose the missing number.", "A09 Trace, Write & Complete",
             "complete-missing-numbers", "missing-number-three-strips-v1",
             {"sequence_1":"visual stepping stones representing 1 2 gap 4 5", "sequence_2":"train carriages representing 6 gap 8 9 10", "sequence_3":"balloon row representing 11 12 13 gap 15"},
             "Read each sequence aloud and pause at the missing place.", cols=1),
    15: spec("Join Groups", "Understand addition by combining.",
             "Count the two groups. Join them and choose the total.", "A07 Count & Record",
             "picture-addition-join-groups", "addition-three-stories-v1",
             {"story_1_left":"two red birds", "story_1_right":"one red bird", "story_2_left":"three blue cars", "story_2_right":"two blue cars", "story_3_left":"four yellow flowers", "story_3_right":"two yellow flowers"},
             "Count each part, sweep both groups together with a finger, then count the total.", cols=2),
    16: spec("Take Away", "Understand subtraction.",
             "Count each set. Cross out the objects taken away and choose how many remain.", "A07 Count & Record",
             "picture-subtraction-cross-out", "subtraction-three-stories-v1",
             {"story_1":"five apples with two clearly separated at the end", "story_2":"six fish with one clearly separated", "story_3":"seven balloons with three clearly separated"},
             "Count the starting set, mark the objects taken away, then count what remains.", cols=3),
    17: spec("Before & After", "Identify numbers before and after.",
             "Look at the centre number. Choose the number before and the number after.", "A08 Compare & Choose",
             "number-before-after", "before-after-four-rows-v1",
             {"row_1":"central numeral 4 token with two blank-side visual anchors", "row_2":"central numeral 7 token with two blank-side visual anchors", "row_3":"central numeral 12 token with two blank-side visual anchors", "row_4":"central numeral 18 token with two blank-side visual anchors"},
             "Say the counting sequence around the centre number, then choose before and after.", cols=1),
    18: spec("Number Order", "Arrange numbers in order.",
             "Put each set of numeral tokens in order from smallest to greatest.", "A04 Sequence & Retell",
             "order-numerals", "number-order-three-rows-v1",
             {"row_1":"mixed numeral tokens 3 1 2", "row_2":"mixed numeral tokens 8 6 7 5", "row_3":"mixed numeral tokens 14 11 13 12"},
             "Find the smallest number first, then continue counting forward.", cols=1),
    19: spec("Number Line", "Read and use a number line.",
             "Follow each jump on the number line and choose where it lands.", "A11 Follow Path & Solve",
             "number-line-jumps", "number-line-three-strips-v1",
             {"line_1":"frog beside a clean 0-to-5 stepping path", "line_2":"rabbit beside a clean 5-to-10 stepping path", "line_3":"bee beside a clean 10-to-15 stepping path"},
             "Point to the starting number and count one spoken number for each jump.", cols=1),
    20: spec("Compare Numbers", "Compare number values.",
             "Look at each number pair. Circle the greater or smaller number as instructed.", "A08 Compare & Choose",
             "compare-numerals", "compare-number-four-pairs-v1",
             {"pair_1":"numeral tokens 4 and 7", "pair_2":"numeral tokens 9 and 6", "pair_3":"numeral tokens 12 and 15", "pair_4":"numeral tokens 18 and 13"},
             "Ask which number comes later when counting; later means greater.", cols=2),
    21: spec("Math Stories", "Solve simple picture stories.",
             "Look at each picture story. Count, think, and choose the answer.", "A11 Follow Path & Solve",
             "picture-maths-stories", "math-story-three-scenes-v1",
             {"story_1":"three ducks in water and two ducks approaching", "story_2":"six biscuits with two being removed", "story_3":"four children sharing four balls one each"},
             "Ask what happened first, what changed, and what must be counted now.", cols=3),
    22: spec("2D Shapes", "Recognise basic 2D shapes.",
             "Name each shape and match it to the object with the same shape.", "A05 Read/Look & Match",
             "match-2d-shape-to-object", "shape-match-four-pairs-v1",
             {"shape_circle":"plain circle", "shape_square":"plain square", "shape_triangle":"plain triangle", "shape_rectangle":"plain rectangle", "object_clock":"round clock", "object_window":"square window", "object_flag":"triangular flag", "object_door":"rectangular door"},
             "Trace the outline in the air before matching each shape.", cols=2),
    23: spec("3D Objects", "Recognise common 3D objects.",
             "Look at each solid object. Match it to the familiar object with the same form.", "A05 Read/Look & Match",
             "match-3d-solid-to-object", "solid-match-four-pairs-v1",
             {"solid_sphere":"plain sphere", "solid_cube":"plain cube", "solid_cylinder":"plain cylinder", "solid_cone":"plain cone", "object_ball":"ball", "object_dice":"dice", "object_can":"food can", "object_party_hat":"party hat"},
             "Ask whether each object rolls, stacks, or has a point before matching.", cols=2),
    24: spec("Shape Hunt", "Find shapes in familiar objects.",
             "Scan the scene. Find and name the circle, square, triangle and rectangle.", "A12 Observe, Find & Name",
             "shape-hunt-scene", "hero-scene-target-strip-v2",
             {"main_scene":"simple playground with round sun, square window, triangular roof and rectangular bench", "target_circle":"plain circle", "target_square":"plain square", "target_triangle":"plain triangle", "target_rectangle":"plain rectangle"},
             "Name one target shape, then let the child find the matching form in the scene.", cols=3),
    25: spec("Patterns", "Recognise repeating patterns.",
             "Study each repeating pattern. Say the repeating unit.", "A04 Sequence & Retell",
             "identify-repeating-pattern", "pattern-three-strips-v1",
             {"pattern_1":"red blue red blue red blue objects", "pattern_2":"circle square circle square circle square", "pattern_3":"leaf flower flower leaf flower flower"},
             "Cover part of the row and ask the child to predict what repeats.", cols=1),
    26: spec("Complete the Pattern", "Complete repeating patterns.",
             "Look at each pattern. Choose the item that comes next.", "A04 Sequence & Retell",
             "complete-repeating-pattern", "complete-pattern-four-rows-v1",
             {"row_1":"apple banana apple banana apple gap", "row_2":"circle triangle triangle circle triangle triangle gap", "row_3":"red block blue block green block red block blue block gap", "row_4":"small star big star small star big star gap"},
             "Say the repeating unit aloud, then choose the next item.", cols=1),
    27: spec("Position Words", "Understand basic positional words.",
             "Look at each picture and choose the correct position word.", "A08 Compare & Choose",
             "position-word-choice", "position-six-scenes-v1",
             {"scene_in":"ball in a box", "scene_on":"cat on a mat", "scene_under":"shoe under a chair", "scene_above":"kite above a tree", "scene_beside":"dog beside a child", "scene_between":"ball between two cones"},
             "Ask the child to point to the reference object before naming the position.", cols=3),
    28: spec("Directions", "Follow simple directions.",
             "Follow each arrow path from the starting object to the destination.", "A11 Follow Path & Solve",
             "follow-direction-path", "directions-three-paths-v1",
             {"path_1":"mouse and cheese with open route space", "path_2":"bee and flower with open route space", "path_3":"car and garage with open route space"},
             "Use left, right, up and down words while tracing each route with a finger.", cols=1),
    29: spec("Big & Small", "Compare size.",
             "Look at each pair. Circle the big or small object as instructed.", "A08 Compare & Choose",
             "compare-big-small", "size-comparison-four-pairs-v1",
             {"pair_1":"large and small elephant", "pair_2":"large and small ball", "pair_3":"large and small leaf", "pair_4":"large and small cup"},
             "Compare matching objects only and keep attention on size, not colour.", cols=2),
    30: spec("Tall & Short", "Compare height.",
             "Look at each pair. Circle the taller or shorter item as instructed.", "A08 Compare & Choose",
             "compare-tall-short", "height-comparison-four-pairs-v1",
             {"pair_1":"tall and short tree", "pair_2":"tall and short bottle", "pair_3":"tall and short child silhouettes", "pair_4":"tall and short tower"},
             "Align the bases mentally or with a finger before comparing height.", cols=2),
    31: spec("Heavy & Light", "Compare weight using familiar clues.",
             "Look at each pair. Circle the object that is likely heavier or lighter.", "A08 Compare & Choose",
             "compare-heavy-light", "weight-comparison-four-pairs-v1",
             {"pair_1":"rock and feather", "pair_2":"watermelon and grape", "pair_3":"full schoolbag and pencil", "pair_4":"metal pan and paper plate"},
             "Invite the child to mime lifting both objects before choosing.", cols=2),
    32: spec("Long & Short", "Compare length.",
             "Look at each pair. Circle the longer or shorter item as instructed.", "A08 Compare & Choose",
             "compare-long-short", "length-comparison-four-pairs-v1",
             {"pair_1":"long and short ribbon aligned at one end", "pair_2":"long and short pencil aligned at one end", "pair_3":"long and short snake aligned at one end", "pair_4":"long and short rope aligned at one end"},
             "Check that both objects begin at the same point before comparing length.", cols=2),
    33: spec("Full & Empty", "Distinguish full and empty containers.",
             "Look at each pair. Circle the full or empty container as instructed.", "A08 Compare & Choose",
             "compare-full-empty", "full-empty-four-pairs-v1",
             {"pair_1":"full and empty glass", "pair_2":"full and empty basket", "pair_3":"full and empty bucket", "pair_4":"full and empty lunchbox"},
             "Ask what is inside each container and whether there is any space left.", cols=2),
    34: spec("Capacity", "Estimate and compare capacity.",
             "Look at each container pair. Choose which can hold more or less.", "A08 Compare & Choose",
             "compare-capacity", "capacity-four-pairs-v1",
             {"pair_1":"small cup and large jug", "pair_2":"small bowl and large pot", "pair_3":"short bottle and large bucket", "pair_4":"small box and large storage tub"},
             "Use the words holds more and holds less; do not confuse height with capacity.", cols=2),
    35: spec("Time Awareness", "Recognise daily routines and their order.",
             "Look at the routine pictures. Put morning, afternoon, evening and night in order.", "A04 Sequence & Retell",
             "daily-routine-order", "daily-routine-four-events-v1",
             {"morning":"child waking with sunrise", "afternoon":"child eating lunch", "evening":"child playing at sunset", "night":"child sleeping with moon"},
             "Use familiar routine language and ask what happens first, next and last.", cols=4),
    36: spec("Sorting", "Sort objects by one attribute.",
             "Look at the model rule. Place each object in the correct group.", "A06 Sort & Classify",
             "sort-by-one-attribute", "classification-grid-v2",
             {"category_red":"one red exemplar object", "category_blue":"one blue exemplar object", "item_1":"red apple", "item_2":"blue car", "item_3":"red ball", "item_4":"blue fish", "item_5":"red flower", "item_6":"blue kite"},
             "Name the sorting rule once, then let the child place every item.", cols=3),
    37: spec("Classifying", "Classify objects by two attributes.",
             "Look at each object. Place it by colour and shape.", "A06 Sort & Classify",
             "classify-two-attributes", "classification-matrix-v1",
             {"red_circle":"red circle token", "red_square":"red square token", "blue_circle":"blue circle token", "blue_square":"blue square token", "yellow_circle":"yellow circle token", "yellow_square":"yellow square token"},
             "Ask the child to name both attributes before placing each object.", cols=3),
    38: spec("Picture Graph", "Read a simple picture graph.",
             "Count the pictures in each category. Answer the comparison questions.", "A08 Compare & Choose",
             "read-picture-graph", "picture-graph-four-categories-v1",
             {"apples":"five isolated apple icons", "bananas":"three isolated banana icons", "grapes":"four isolated grape icons", "oranges":"two isolated orange icons"},
             "Count one category at a time and compare the category totals.", cols=4),
    39: spec("Problem Solving", "Apply maths thinking.",
             "Look at each visual problem. Choose the answer and explain how you know.", "A11 Follow Path & Solve",
             "mixed-maths-problems", "problem-solving-three-cards-v1",
             {"problem_1":"five birds with two flying away", "problem_2":"three red blocks and three blue blocks to compare", "problem_3":"simple AB pattern ending with a gap"},
             "Ask what information matters, then let the child show the reasoning with a finger.", cols=3),
    40: spec("Maths Around Me", "Recognise maths in daily life.",
             "Find and name the numbers, shapes, sizes and patterns in the scene.", "A12 Observe, Find & Name",
             "maths-around-me-find", "hero-scene-target-strip-v2",
             {"main_scene":"simple market scene containing price numerals, round fruit, rectangular boxes, big and small baskets, and a striped pattern", "target_number":"isolated numeral cue", "target_shape":"isolated circle cue", "target_size":"isolated big-small cue", "target_pattern":"isolated stripe pattern cue"},
             "Give one target at a time and let the child scan, point and name it.", cols=3),
    41: spec("Maths Review", "Review key maths skills from the book.",
             "Complete one short task for number, shape, pattern and comparison.", "A14 Mixed Review",
             "mixed-maths-review", "maths-review-four-quadrants-v1",
             {"number_task":"group of six stars for counting", "shape_task":"circle triangle square and rectangle", "pattern_task":"red blue red blue gap", "comparison_task":"large ball and small ball"},
             "Complete one quadrant at a time and praise the strategy, not speed.", cols=2),
    42: spec("Certificate", "Celebrate completion of Early Maths Adventures.",
             "Celebrate the child's effort and completion.", "A15 Celebrate & Certify",
             "certificate-celebration-assets", "certificate-celebration-v1",
             {"celebration_cluster":"small confetti and maths-symbol celebration cluster", "badge":"plain gold achievement badge without text", "star_border":"small decorative stars kept separate from writing areas"},
             "Read the completed certificate aloud and celebrate effort.", cols=3, response="The publishing engine adds all certificate wording and name/date lines."),
    43: spec("I Am a Maths Explorer", "Reflect on learning and confidence.",
             "Choose the maths skills you enjoyed and say one thing you can do now.", "A10 Speak, Listen & Respond",
             "maths-reflection-choice", "hero-plus-choice-assets-v2",
             {"hero_scene":"one proud child holding a completed maths workbook", "choice_numbers":"small number activity sample", "choice_shapes":"small shape activity sample", "choice_patterns":"small pattern activity sample"},
             "Invite the child to choose one skill and complete the sentence: I can ___.", cols=2),
    44: spec("Early Maths Adventures — Back Cover", "Provide a clean series back-cover illustration.",
             "No child activity on this publishing page.", "A00 Publishing Page",
             "back-cover-illustration", "back-cover-single-scene-v1",
             {"back_cover_scene":"small friendly maths discovery scene with blocks, shapes and counting objects, no text or logo"},
             "No teacher cue is printed on the back cover.", cols=1, response="No worksheet mechanics or response controls."),
}


def illustration_prompt(page_id: str, s: dict[str, Any]) -> str:
    assets = "\n".join(f"- {name}: {description}" for name, description in s["assets"].items())
    names = ", ".join(s["assets"])
    return f"""BCube Phase 2 Illustration Asset Prompt — {page_id}

BOOK AND LEVEL
Book: {BOOK}
Level: {LEVEL} (4+)
Exact page title: {s['title']}
Learning objective: {s['objective']}
Primary child action: {s['instruction']}
Assigned page archetype: {s['archetype']}
Primary mechanic: {s['mechanic']}
Asset layout standard: {s['layout']}

ILLUSTRATION ROLE — LOCKED
Create one illustration-only source sheet containing exactly the named assets below. The publishing engine will crop these assets and add all text, numerals, response controls, lines, circles, arrows, bins, sequence slots, answer choices, branding, teacher cue and page number.

EXACT NAMED ASSETS
{assets}

CROP AND SPACING LOCK
Use the exact crop names: {names}. Keep every named asset entirely inside its assigned visual zone, isolated on pure white, with wide gutters and no merged shadows. Do not add or remove assets. Do not combine neighbouring assets into one scene unless the asset description explicitly says scene or pair.

STYLE
Premium commercial preschool-publishing illustration for age 4+, large recognisable forms, clean rounded outlines, refined flat colour, correct anatomy where applicable, minimal detail, no photorealism.

STRICT EXCLUSIONS
No complete workbook page, logo, mascot, title, instruction, label, generated words, generated numerals unless a numeral token is explicitly named, border, panel, card, response box, answer state, matching line, arrow, check mark, crossed-out answer, completed sequence, teacher, classroom group, watermark or QR code.

ALIGNMENT GATE
Reject the illustration if any expected asset is missing, renamed, duplicated, touching another crop zone, semantically different from its description, or unable to support the mechanic “{s['mechanic']}”."""


def execution_prompt(page_id: str, s: dict[str, Any]) -> str:
    crops = json.dumps(s["crops"], separators=(",", ":"))
    return f"""BCube Runtime Page Execution Contract — {page_id}

IDENTITY
Book: {BOOK}
Level: {LEVEL}
Title: {s['title']}
Objective: {s['objective']}

CHILD ACTION
{s['instruction']}

ARCHETYPE AND MECHANIC
Archetype: {s['archetype']}
Mechanic: {s['mechanic']}
Layout: {s['layout']}
Response purpose: {s['response']}

ILLUSTRATION CONTRACT
Expected source filename: {page_id}.png
Required named assets: {', '.join(s['assets'])}
Crop manifest: {crops}
The renderer must use only these named crops. It must not infer substitute assets or use a generic response panel.

TEACHER CUE
{s['teacher']}

LOCKED RULES
- Illustration and page contract must use the same page ID, asset names, asset meanings and crop manifest.
- No Parent Prompt, Home Connection or homework panel.
- No generic lower response box and no unexplained circles, lines, cards or frames.
- Every visible response control must belong to the mechanic “{s['mechanic']}”.
- Missing or invalid assets fail closed; allow_fallback is false.
- Typography and worksheet mechanics are added deterministically, never generated inside the artwork."""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--output", type=Path)
    ap.add_argument("--sheet", default="All Page Prompts")
    args = ap.parse_args()

    source = args.workbook.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Workbook not found: {source}")
    output = (args.output or source.with_name(source.stem + "_EM_LKG_FULL_ALIGNED.xlsx")).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    if output == source:
        backup = source.with_suffix(source.suffix + ".bak")
        shutil.copy2(source, backup)

    wb = load_workbook(source)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]
    headers = {str(cell.value).strip(): i + 1 for i, cell in enumerate(ws[1]) if cell.value is not None}
    required = ["Prompt ID", "Level", "Book Slug", "Physical Page", "Page Title", "Learning Objective",
                "Complete Standalone Illustration Prompt", "Output Filename", "Status",
                "Phase 2 Page Execution Prompt", "Phase 2 Archetype", "Phase 2 Parent Box",
                "Phase 2 Patch Status", "Crop-Based Asset Sheet", "Asset Layout Standard",
                "Crop-Safe Spacing Rule", "Crop Manifest JSON", "Ready to Test", "Prompt Validation Status"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing workbook columns: {missing}")

    updated: list[str] = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers["Level"]).value != LEVEL or ws.cell(row, headers["Book Slug"]).value != SLUG:
            continue
        page = int(ws.cell(row, headers["Physical Page"]).value)
        s = SPECS.get(page)
        if s is None:
            continue
        page_id = str(ws.cell(row, headers["Prompt ID"]).value)
        expected = f"EM-LKG-V4-P{page:03d}"
        if page_id != expected:
            raise ValueError(f"Row {row}: page ID {page_id!r} does not match {expected!r}")
        ws.cell(row, headers["Page Title"]).value = s["title"]
        ws.cell(row, headers["Learning Objective"]).value = s["objective"]
        ws.cell(row, headers["Complete Standalone Illustration Prompt"]).value = illustration_prompt(page_id, s)
        ws.cell(row, headers["Output Filename"]).value = f"{page_id}.png"
        ws.cell(row, headers["Status"]).value = "EM LKG full illustration-contract alignment"
        ws.cell(row, headers["Phase 2 Page Execution Prompt"]).value = execution_prompt(page_id, s)
        ws.cell(row, headers["Phase 2 Archetype"]).value = s["archetype"]
        ws.cell(row, headers["Phase 2 Parent Box"]).value = "REMOVE"
        ws.cell(row, headers["Phase 2 Patch Status"]).value = "FULLY ALIGNED — illustration + runtime contract"
        ws.cell(row, headers["Crop-Based Asset Sheet"]).value = "YES"
        ws.cell(row, headers["Asset Layout Standard"]).value = s["layout"]
        ws.cell(row, headers["Crop-Safe Spacing Rule"]).value = "Keep every named asset fully inside its crop zone with wide white gutters; no touching, overlap, merged shadows or neighbouring-object contamination."
        ws.cell(row, headers["Crop Manifest JSON"]).value = json.dumps(s["crops"], separators=(",", ":"))
        ws.cell(row, headers["Ready to Test"]).value = "PROMPT UPDATED"
        ws.cell(row, headers["Prompt Validation Status"]).value = "ALIGNED — TEST REQUIRED"
        updated.append(page_id)

    if len(updated) != len(SPECS):
        missing_ids = sorted(set(f"EM-LKG-V4-P{p:03d}" for p in SPECS) - set(updated))
        raise ValueError(f"Expected {len(SPECS)} pages but updated {len(updated)}; missing {missing_ids}")
    wb.save(output)
    print(json.dumps({
        "output": str(output),
        "updated_pages": len(updated),
        "first_page": updated[0],
        "last_page": updated[-1],
        "policy": "Illustration prompt, asset names, crop manifest, mechanic and runtime execution contract are updated together."
    }, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"BCube Early Maths LKG illustration rebuild FAIL: {exc}")
        raise SystemExit(2)
