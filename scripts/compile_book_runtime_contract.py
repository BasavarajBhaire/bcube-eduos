#!/usr/bin/env python3
"""Compile one BCube book into one self-contained runtime-contract JSON.

The spreadsheet row is the authoritative bridge between illustration generation
and page composition. Every READY runtime page stores the exact illustration
prompt, crop manifest, layout standard and their hashes. A page is BLOCKED when
that evidence is missing or inconsistent; no generic fallback is created.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REQUIRED_WORKBOOK_COLUMNS = {
    "Prompt ID", "Level", "Book Slug", "Physical Page", "Printed Page",
    "Page Type", "Page Title", "Learning Objective", "Activity Type",
    "Complete Standalone Illustration Prompt", "Output Filename",
    "Phase 2 Page Execution Prompt", "Phase 2 Archetype",
    "Asset Layout Standard", "Crop Manifest JSON",
}


class CompileError(RuntimeError):
    pass


def load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise CompileError(f"{path} must contain a JSON object")
    return value


def clean(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def normalise_level(value: str) -> str:
    key = value.strip().lower()
    mapping = {"nursery": "Nursery", "lkg": "LKG", "ukg": "UKG"}
    if key not in mapping:
        raise CompileError(f"Unknown level: {value}")
    return mapping[key]


def page_number_from_id(page_id: str) -> int:
    match = re.search(r"-P(\d{3})$", page_id)
    if not match:
        raise CompileError(f"Cannot read physical page from {page_id}")
    return int(match.group(1))


def read_workbook_rows(path: Path, sheet_name: str) -> tuple[dict[str, dict[str, Any]], str]:
    if not path.is_file():
        raise CompileError(f"Workbook not found: {path}")
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise CompileError(f"Worksheet {sheet_name!r} not found in {path}")
    sheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_WORKBOOK_COLUMNS - set(headers)
    if missing:
        raise CompileError(f"Workbook is missing columns: {sorted(missing)}")
    rows: dict[str, dict[str, Any]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        page_id = clean(row.get("Prompt ID"))
        if page_id:
            rows[page_id] = row
    return rows, hashlib.sha256(path.read_bytes()).hexdigest()


def parse_crop_manifest(raw: Any, page_id: str) -> dict[str, list[float]]:
    if isinstance(raw, dict):
        value = raw
    else:
        text = clean(raw)
        if not text:
            raise CompileError(f"{page_id}: Crop Manifest JSON is empty")
        try:
            value = json.loads(text)
        except json.JSONDecodeError as exc:
            raise CompileError(f"{page_id}: invalid Crop Manifest JSON: {exc}") from exc
    if "asset_crops" in value and isinstance(value["asset_crops"], dict):
        value = value["asset_crops"]
    if "crops" in value and isinstance(value["crops"], dict):
        value = value["crops"]
    if not isinstance(value, dict) or not value:
        raise CompileError(f"{page_id}: crop manifest must contain named crops")
    result: dict[str, list[float]] = {}
    for name, coords in value.items():
        if isinstance(coords, dict):
            coords = [coords.get("x0"), coords.get("y0"), coords.get("x1"), coords.get("y1")]
        if not isinstance(coords, (list, tuple)) or len(coords) != 4:
            raise CompileError(f"{page_id}: crop {name!r} must contain four coordinates")
        numbers = [float(v) for v in coords]
        if not all(0 <= v <= 1 for v in numbers) or not (numbers[0] < numbers[2] and numbers[1] < numbers[3]):
            raise CompileError(f"{page_id}: crop {name!r} is outside normalised bounds")
        result[clean(name)] = numbers
    return result


def mechanic_from(row: dict[str, Any]) -> tuple[str, str]:
    archetype = clean(row.get("Phase 2 Archetype"))
    activity = clean(row.get("Activity Type")).casefold()
    title = clean(row.get("Page Title")).casefold()
    text = f"{archetype} {activity} {title}".casefold()
    rules = [
        (("count" in text and "circle" in text), ("count-and-circle-number", "circle-number")),
        (("count" in text and ("write" in text or "record" in text)), ("count-and-record", "write-number")),
        (("match" in text), ("match-pairs", "draw-line")),
        (("sort" in text or "classif" in text), ("sort-and-classify", "place-or-write")),
        (("more" in text or "fewer" in text or "less" in text or "compare" in text), ("compare-quantities", "circle-choice")),
        (("sequence" in text or "before" in text or "after" in text or "missing number" in text), ("number-sequence", "write-or-circle")),
        (("trace" in text), ("trace-and-copy", "trace-write")),
        (("shape" in text and "pattern" not in text), ("shape-identification", "circle-or-match")),
        (("pattern" in text), ("pattern-completion", "choose-or-draw")),
        (("position" in text or "above" in text or "below" in text or "inside" in text or "outside" in text), ("position-identification", "circle-choice")),
        (("add" in text or "addition" in text), ("picture-addition", "write-number")),
        (("subtract" in text or "take away" in text), ("picture-subtraction", "write-number")),
        (("speak" in text or "say" in text or "respond" in text), ("picture-supported-oral-response", "oral")),
        (("find" in text or "observe" in text or "name" in text), ("observe-find-name", "point-or-circle")),
    ]
    for condition, result in rules:
        if condition:
            return result
    return "page-specific-activity", clean(row.get("Activity Type")) or "page-specific-response"


def build_ready_page(source: dict[str, Any], source_path: Path, row: dict[str, Any], workbook_path: Path, workbook_sha: str) -> dict[str, Any]:
    page_id = clean(source.get("prompt_id"))
    page = source.get("page", {})
    curriculum = source.get("curriculum", {})
    illustration_prompt = clean(row.get("Complete Standalone Illustration Prompt"))
    execution_prompt = clean(row.get("Phase 2 Page Execution Prompt"))
    output_filename = clean(row.get("Output Filename")) or f"{page_id}.png"
    layout_standard = clean(row.get("Asset Layout Standard"))
    archetype = clean(row.get("Phase 2 Archetype"))
    objective = clean(row.get("Learning Objective")) or clean(curriculum.get("objective"))
    if not all((illustration_prompt, execution_prompt, output_filename, layout_standard, archetype, objective)):
        raise CompileError(f"{page_id}: required illustration/runtime alignment evidence is missing")
    crops = parse_crop_manifest(row.get("Crop Manifest JSON"), page_id)
    mechanic, response_mode = mechanic_from(row)
    assets = list(crops)
    physical = int(row.get("Physical Page") or page.get("physical") or page_number_from_id(page_id))
    printed_raw = row.get("Printed Page") if row.get("Printed Page") is not None else page.get("printed")
    printed = int(printed_raw) if str(printed_raw or "").strip().isdigit() else None
    instruction = execution_prompt
    teacher_cue = clean(curriculum.get("teacher_facilitation")) or "Model one example, then invite the child to complete the remaining activity independently."
    return {
        "identity": {
            "page_id": page_id,
            "physical_page": physical,
            "printed_page": printed,
            "title": clean(row.get("Page Title")) or clean(page.get("title")),
            "page_type": clean(row.get("Page Type")) or clean(page.get("type")) or "learning_page",
        },
        "learning": {
            "objective": objective,
            "instruction": instruction,
            "expected_response": clean(curriculum.get("evidence")) or f"The child completes the exact {clean(row.get('Activity Type')).lower()} response.",
            "model_text": None,
        },
        "activity": {
            "archetype": archetype,
            "mechanic": mechanic,
            "response_mode": response_mode,
            "activity_count": max(1, len(assets)),
        },
        "illustration": {
            "source_asset": output_filename,
            "artwork_only": True,
            "assets": assets,
            "asset_crops": crops,
            "prompt": illustration_prompt,
            "prompt_sha256": sha256_text(illustration_prompt),
        },
        "mechanics": {
            "page_execution_prompt": execution_prompt,
            "asset_order": assets,
            "activity_type": clean(row.get("Activity Type")),
            "crop_manifest_sha256": sha256_text(json.dumps(crops, sort_keys=True, separators=(",", ":"))),
        },
        "guidance": {"teacher_cue": teacher_cue},
        "layout": {
            "template": layout_standard,
            "parent_panel": False,
            "home_connection": False,
            "generic_response_panel": False,
        },
        "validation": {
            "allow_fallback": False,
            "status": "READY",
            "required_asset_count": len(assets),
            "required_response_count": max(1, len(assets)),
            "blocking_reasons": [],
            "illustration_contract_aligned": True,
        },
        "source_lineage": {
            "source_prompt_id": page_id,
            "source_file": source_path.relative_to(ROOT).as_posix(),
            "workbook_file": str(workbook_path),
            "workbook_sha256": workbook_sha,
            "illustration_prompt_sha256": sha256_text(illustration_prompt),
            "execution_prompt_sha256": sha256_text(execution_prompt),
            "compiler": "compile_book_runtime_contract.py",
        },
    }


def blocked_page(source: dict[str, Any], source_path: Path, reason: str) -> dict[str, Any]:
    page_id = clean(source.get("prompt_id"))
    page = source.get("page", {})
    curriculum = source.get("curriculum", {})
    return {
        "identity": {"page_id": page_id, "physical_page": int(page.get("physical") or page_number_from_id(page_id)), "printed_page": page.get("printed"), "title": clean(page.get("title")), "page_type": clean(page.get("type")) or "learning_page"},
        "learning": {"objective": clean(curriculum.get("objective")), "instruction": clean(curriculum.get("instruction")), "expected_response": clean(curriculum.get("evidence")), "model_text": None},
        "activity": {"archetype": "CONTRACT_REQUIRED", "mechanic": "CONTRACT_REQUIRED", "response_mode": "CONTRACT_REQUIRED", "activity_count": 1},
        "illustration": {"source_asset": f"{page_id}.png", "artwork_only": True, "assets": ["CONTRACT_REQUIRED"], "asset_crops": {"CONTRACT_REQUIRED": [0.0, 0.0, 1.0, 1.0]}, "prompt": "", "prompt_sha256": ""},
        "mechanics": {"contract_required": True},
        "guidance": {"teacher_cue": clean(curriculum.get("teacher_facilitation")) or "Page-specific teacher cue required."},
        "layout": {"template": "CONTRACT_REQUIRED", "parent_panel": False, "home_connection": False, "generic_response_panel": False},
        "validation": {"allow_fallback": False, "status": "BLOCKED", "required_asset_count": 1, "required_response_count": 1, "blocking_reasons": [reason], "illustration_contract_aligned": False},
        "source_lineage": {"source_prompt_id": page_id, "source_file": source_path.relative_to(ROOT).as_posix(), "compiler": "compile_book_runtime_contract.py"},
    }


def should_compile_page(source: dict[str, Any], minimum_physical_page: int) -> bool:
    page = source.get("page", {})
    physical = int(page.get("physical") or 0)
    page_type = clean(page.get("type")).casefold()
    return physical >= minimum_physical_page and page_type not in {"cover", "copyright", "contents", "welcome", "front_matter"}


def compile_book(args: argparse.Namespace) -> dict[str, Any]:
    level_key = args.level.strip().lower()
    level = normalise_level(args.level)
    source_dir = ROOT / "production-prompts" / args.book / level_key / "v4" / "pages"
    if not source_dir.is_dir():
        raise CompileError(f"Source page directory not found: {source_dir}")
    workbook_path = args.workbook.expanduser().resolve()
    workbook_rows, workbook_sha = read_workbook_rows(workbook_path, args.sheet)
    output = ROOT / "runtime-contracts" / level_key / f"{args.book}.json"
    source_files = sorted(source_dir.glob("*.json"))
    pages: dict[str, Any] = {}
    book_title = ""
    prefix = ""
    for path in source_files:
        source = load_json(path)
        if not should_compile_page(source, args.minimum_physical_page):
            continue
        page_id = clean(source.get("prompt_id"))
        book = source.get("book", {})
        book_title = book_title or clean(book.get("name"))
        prefix = prefix or clean(book.get("prefix"))
        row = workbook_rows.get(page_id)
        if not row:
            pages[page_id] = blocked_page(source, path, "No matching illustration-prompt spreadsheet row exists for this page ID.")
            continue
        if clean(row.get("Level")).casefold() != level.casefold() or clean(row.get("Book Slug")) != args.book:
            pages[page_id] = blocked_page(source, path, "Spreadsheet row identity does not match the requested level and book.")
            continue
        try:
            pages[page_id] = build_ready_page(source, path, row, workbook_path, workbook_sha)
        except CompileError as exc:
            pages[page_id] = blocked_page(source, path, str(exc))
    ordered = dict(sorted(pages.items(), key=lambda item: item[1]["identity"]["physical_page"]))
    ready = sum(p["validation"]["status"] == "READY" for p in ordered.values())
    blocked = len(ordered) - ready
    result = {
        "contract_version": "bcube-book-runtime-contract-v1",
        "book": {"title": book_title or args.book, "slug": args.book, "prefix": prefix},
        "level": level,
        "age": {"Nursery": "3+", "LKG": "4+", "UKG": "5+"}[level],
        "allow_fallback": False,
        "illustration_alignment": {"workbook": str(workbook_path), "workbook_sha256": workbook_sha, "sheet": args.sheet, "policy": "Page contract assets, crop manifest, layout and execution prompt must come from the same spreadsheet row as the illustration prompt."},
        "pages": ordered,
        "compile_summary": {"source_page_count": len(ordered), "runtime_page_count": len(ordered), "ready_pages": ready, "blocked_pages": blocked, "illustration_aligned_pages": sum(bool(p["validation"].get("illustration_contract_aligned")) for p in ordered.values()), "policy": "BLOCKED pages do not render; no generic fallback is permitted."},
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return {"output": str(output), **result["compile_summary"]}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compile one illustration-aligned BCube book runtime contract")
    parser.add_argument("--level", choices=["nursery", "lkg", "ukg"], required=True)
    parser.add_argument("--book", required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--sheet", default="All Page Prompts")
    parser.add_argument("--minimum-physical-page", type=int, default=8)
    return parser.parse_args()


def main() -> int:
    print(json.dumps(compile_book(parse_args()), indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (CompileError, OSError, json.JSONDecodeError, ValueError) as exc:
        print(f"BCube book runtime compile FAIL: {exc}")
        raise SystemExit(2)
