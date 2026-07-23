#!/usr/bin/env python3
"""Generate a bounded BCube illustration preview before a paid full batch.

The command copies only the first N matching workbook rows into a temporary
workbook and delegates generation to ``bcube_cloud_illustrations.py``. This
provides a hard request-count gate: the provider never receives more than the
requested preview size.
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import List, Sequence

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CLOUD_SCRIPT = ROOT / "scripts" / "bcube_cloud_illustrations.py"


def filtered_rows(workbook: Path, sheet_name: str, level: str) -> tuple[list[str], list[list[object]]]:
    if not workbook.is_file():
        raise FileNotFoundError(workbook)
    source = load_workbook(workbook, read_only=True, data_only=True)
    try:
        if sheet_name not in source.sheetnames:
            raise ValueError("Worksheet %r not found" % sheet_name)
        sheet = source[sheet_name]
        values = list(sheet.iter_rows(values_only=True))
    finally:
        source.close()
    if not values:
        raise ValueError("Workbook sheet is empty")
    headers = [str(value or "").strip() for value in values[0]]
    try:
        level_index = headers.index("Level")
    except ValueError as exc:
        raise ValueError("Workbook must contain a Level column") from exc
    rows: list[list[object]] = []
    for raw in values[1:]:
        row = list(raw)
        row_level = str(row[level_index] or "").strip().lower() if level_index < len(row) else ""
        if level == "all" or row_level == level:
            rows.append(row)
    return headers, rows


def write_preview_workbook(source_path: Path, sheet_name: str, headers: Sequence[str], rows: Sequence[Sequence[object]], target: Path) -> None:
    source = load_workbook(source_path)
    try:
        sheet = source[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(list(headers))
        for row in rows:
            sheet.append(list(row))
        source.save(target)
    finally:
        source.close()


def build_delegate_command(args: argparse.Namespace, preview_workbook: Path, batch_name: str) -> List[str]:
    command = [
        sys.executable,
        str(CLOUD_SCRIPT),
        "--workbook",
        str(preview_workbook),
        "--sheet",
        args.sheet,
        "--level",
        "all",
        "--provider",
        args.provider,
        "--workers",
        str(args.workers),
        "--timeout",
        str(args.timeout),
        "--poll-interval",
        str(args.poll_interval),
        "--cost-per-second",
        str(args.cost_per_second),
        "--max-retries",
        str(args.max_retries),
        "--min-occupancy",
        str(args.min_occupancy),
        "--output-root",
        str(args.output_root),
        "--batch-name",
        batch_name,
        "--require-complete",
    ]
    if args.resume:
        command.append("--resume")
    if args.endpoint_id:
        command.extend(["--endpoint-id", args.endpoint_id])
    if args.api_key:
        command.extend(["--api-key", args.api_key])
    if args.workflow:
        command.extend(["--workflow", str(args.workflow)])
    return command


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate only a small review batch before paid production")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--sheet", default="Complete Book List")
    parser.add_argument("--level", choices=["all", "nursery", "lkg", "ukg"], default="nursery")
    parser.add_argument("--preview", type=int, default=1, help="Maximum prompts sent to the provider")
    parser.add_argument("--provider", choices=["runpod", "mock"], default="mock")
    parser.add_argument("--endpoint-id")
    parser.add_argument("--api-key")
    parser.add_argument("--workflow", type=Path)
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--timeout", type=int, default=900)
    parser.add_argument("--poll-interval", type=float, default=2.0)
    parser.add_argument("--cost-per-second", type=float, default=0.0)
    parser.add_argument("--max-retries", type=int, default=0, help="Defaults to zero for strict cost control")
    parser.add_argument("--min-occupancy", type=float, default=0.55)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--output-root", type=Path, default=ROOT / "production-renders" / "cloud-illustration-batches")
    parser.add_argument("--batch-name")
    args = parser.parse_args()

    if args.preview < 1:
        raise ValueError("--preview must be at least 1")
    if args.workers < 1:
        raise ValueError("--workers must be at least 1")
    if args.provider == "runpod" and (not args.endpoint_id or not args.api_key):
        raise ValueError("RunPod preview requires --endpoint-id and --api-key")

    headers, available = filtered_rows(args.workbook, args.sheet, args.level)
    if not available:
        raise ValueError("No workbook rows matched the selected level")
    selected = available[: args.preview]
    name = args.batch_name or "%s-%s-preview-%d" % (args.workbook.stem.replace(" ", "_"), args.level, len(selected))
    batch_root = args.output_root / name
    batch_root.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="bcube-preview-") as temp_dir:
        preview_workbook = Path(temp_dir) / "bcube-preview.xlsx"
        write_preview_workbook(args.workbook, args.sheet, headers, selected, preview_workbook)
        command = build_delegate_command(args, preview_workbook, name)
        completed = subprocess.run(command, cwd=ROOT)

    preview_manifest = {
        "mode": "PREVIEW",
        "source_workbook": str(args.workbook),
        "level": args.level,
        "available_matching_rows": len(available),
        "preview_limit": args.preview,
        "selected": len(selected),
        "provider": args.provider,
        "max_retries": args.max_retries,
        "full_run_authorized": False,
        "next_step": "Review illustration-review.html. Run bcube_cloud_illustrations.py separately only after explicit approval.",
        "delegate_exit_code": completed.returncode,
    }
    manifest_path = batch_root / "preview-gate.json"
    manifest_path.write_text(json.dumps(preview_manifest, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "status": "PASS" if completed.returncode == 0 else "FAIL",
        "mode": "PREVIEW",
        "selected": len(selected),
        "available_matching_rows": len(available),
        "batch_root": str(batch_root),
        "preview_gate": str(manifest_path),
        "full_run_authorized": False,
    }, indent=2))
    return completed.returncode


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (ValueError, FileNotFoundError, subprocess.CalledProcessError) as exc:
        print("BCube illustration preview ERROR: %s" % exc, file=sys.stderr)
        raise SystemExit(2)
