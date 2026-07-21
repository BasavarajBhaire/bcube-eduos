#!/usr/bin/env python3
"""Generate the complete Curiosity Explorers Nursery Phase 3 package.

The generator is deterministic: the same source data produces the same Markdown,
JSON, CSV and XLSX outputs. It creates documentation only; it never generates or
approves artwork.
"""
from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from textwrap import dedent

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("BCube_Gold_Production_v4/books/nursery/curiosity-explorers")
DOCS = ROOT / "docs"
PROMPTS = ROOT / "production-prompts"
QA = ROOT / "qa"
PUBLISHING = ROOT / "publishing"
WORKBOOK = ROOT / "Curiosity_Explorers_Gold_Production_Prompts_v5.xlsx"

PAGES = [
    (1, "Cover", "Front Matter", "Introduce Curiosity Explorers and the BCube discovery journey.", "Observe, Explore, Question", "Cover composition", "A unique explorer classroom scene with Star and diverse children using magnifying glasses to observe leaves, butterflies, shells and simple science objects."),
    (2, "Copyright", "Front Matter", "Present complete publisher and edition information.", "Print awareness", "Read with an adult", "Clean legal-information page with small approved brand elements and no decorative clutter."),
    (3, "Welcome to Curiosity Explorers", "Front Matter", "Welcome the child into an observation-and-discovery journey.", "Curiosity, Communication", "Point and say", "Star welcomes diverse preschool children into a bright discovery corner with magnifying glass, leaf, shell and butterfly."),
    (4, "How to Use This Book", "Front Matter", "Show adults and children the recurring activity symbols and interaction flow.", "Independence, Communication", "Look, point, talk", "Four large visual steps: look, think, explore and talk, each represented by one clear icon and a child action."),
    (5, "Meet Star the Explorer", "Front Matter", "Introduce Star as a consistent inquiry companion.", "Observe, Question, Explore", "Listen and repeat", "Approved Star mascot holding a magnifying glass with four simple fact callouts: looks carefully, asks questions, explores safely, shares discoveries."),
    (6, "My Explorer Promise", "Front Matter", "Build safe and positive exploration habits.", "Confidence, Responsibility", "Trace and repeat", "Four promise panels with a child observing, asking, exploring safely and caring for nature; include large traceable key words."),
    (7, "Look Carefully", "Module 1 – Observe and Compare", "Find named objects by scanning a simple scene carefully.", "Observation, Vocabulary", "Find and circle", "A simple playground scene containing a butterfly, ball, bird, kite, flower, bench and tree; Star observes from the side with a magnifying glass."),
    (8, "Big and Small", "Module 1 – Observe and Compare", "Compare familiar objects by size.", "Compare, Observe", "Circle the bigger object", "Four pairs of familiar objects with obvious size contrast: balls, teddy bears, cups and leaves."),
    (9, "Same and Different", "Module 1 – Observe and Compare", "Identify one object that differs from a small group.", "Compare, Classify", "Find the different one", "Three rows of four large objects; one item in each row differs by colour, shape or orientation."),
    (10, "Find My Match", "Module 1 – Observe and Compare", "Match identical objects using visual discrimination.", "Observe, Match", "Draw matching lines", "Two columns of six large familiar objects in shuffled order with generous line-drawing space."),
    (11, "What Is Missing?", "Module 1 – Observe and Compare", "Recall and identify a missing object in a simple sequence.", "Memory, Observe", "Choose the missing object", "Three short visual sequences with one blank and two large answer choices beneath each."),
    (12, "Spot the Difference", "Module 1 – Observe and Compare", "Notice small changes between two related scenes.", "Observe, Compare", "Circle three differences", "Two side-by-side preschool garden scenes with exactly three obvious differences and no tiny details."),
    (13, "Look and Count", "Module 1 – Observe and Compare", "Count visible objects from one to five.", "Observe, Early Numeracy", "Count and circle", "Five groups of large nature objects, quantities one through five, with large numeral choices."),
    (14, "Observation Challenge", "Module 1 – Observe and Compare", "Review size, matching, difference and counting skills.", "Observe, Compare, Classify", "Mixed challenge", "Four clearly separated mini-activities using familiar objects from pages 7–13."),
    (15, "Beautiful Leaves", "Module 2 – Nature Explorers", "Notice differences in leaf shape, size and edge.", "Observe, Nature Awareness", "Match and colour", "Four large leaf types with distinct silhouettes and one simple tree reference."),
    (16, "Wonderful Flowers", "Module 2 – Nature Explorers", "Observe flower colour and petal count.", "Observe, Compare", "Count and colour petals", "Four large simple flowers with clearly separated petals and friendly garden accents."),
    (17, "Trees Around Me", "Module 2 – Nature Explorers", "Identify the main visible parts of a tree.", "Observe, Vocabulary", "Point and match", "One large child-friendly tree showing roots, trunk, branches and leaves with four label cards."),
    (18, "Fruits We Eat", "Module 2 – Nature Explorers", "Recognize familiar fruits and connect them with colours.", "Classify, Vocabulary", "Match fruit to colour", "Apple, banana, orange, grapes and watermelon shown as large isolated objects with colour swatches."),
    (19, "Healthy Vegetables", "Module 2 – Nature Explorers", "Recognize familiar vegetables and sort them into a basket.", "Classify, Healthy Habits", "Circle and sort", "Carrot, tomato, potato, brinjal, peas and cucumber arranged around one large basket."),
    (20, "Birds Around Us", "Module 2 – Nature Explorers", "Observe common bird features and identify a nest.", "Observe, Nature Awareness", "Find and match", "Three friendly birds, one nest, feathers and eggs in a simple outdoor setting."),
    (21, "Butterflies and Bees", "Module 2 – Nature Explorers", "Compare two familiar garden insects.", "Compare, Observe", "Tick what you notice", "One large butterfly and one large bee with simple comparison icons for wings, colours and movement."),
    (22, "Weather Today", "Module 2 – Nature Explorers", "Recognize sunny, rainy, cloudy and windy weather.", "Observe, Vocabulary", "Match weather to scene", "Four equal weather panels with one child dressed appropriately in each."),
    (23, "Farm Animals", "Module 3 – Animal Explorers", "Recognize common farm animals and their sounds.", "Classify, Listening", "Match animal to sound", "Cow, sheep, horse, hen and duck as large friendly illustrations with sound cards."),
    (24, "Wild Animals", "Module 3 – Animal Explorers", "Distinguish wild animals from familiar domestic animals.", "Classify, Observe", "Circle wild animals", "Lion, elephant, tiger and giraffe mixed with dog and cat, all in separate clear frames."),
    (25, "Sea Animals", "Module 3 – Animal Explorers", "Identify animals that live in water.", "Classify, Nature Awareness", "Find water animals", "Fish, dolphin, octopus, turtle and starfish mixed with two land animals."),
    (26, "Birds Can Fly", "Module 3 – Animal Explorers", "Recognize common birds and connect them to the sky.", "Classify, Observe", "Match birds to sky", "Parrot, pigeon, crow and sparrow with one clear sky area and simple matching paths."),
    (27, "Animal Homes", "Module 3 – Animal Explorers", "Match common animals with their homes.", "Match, Classify", "Draw matching lines", "Dog and kennel, bird and nest, rabbit and burrow, bee and hive in two aligned columns."),
    (28, "Baby Animals", "Module 3 – Animal Explorers", "Match animal babies with their parents.", "Match, Vocabulary", "Match baby to parent", "Cow-calf, dog-puppy, cat-kitten and hen-chick as four clear pairs."),
    (29, "Animal Food", "Module 3 – Animal Explorers", "Connect familiar animals with foods they eat.", "Match, Classify", "Match animal to food", "Rabbit-carrot, monkey-banana, cow-grass and bird-seeds with large recognizable objects."),
    (30, "Animal Explorer Challenge", "Module 3 – Animal Explorers", "Review animal groups, homes, babies and foods.", "Classify, Recall", "Mixed animal challenge", "Four mini-activities using only animals introduced on pages 23–29."),
    (31, "My Home", "Module 4 – Explore My World", "Recognize familiar rooms and household objects.", "Observe, Vocabulary", "Match object to room", "Four simple room panels: bedroom, kitchen, bathroom and living room, with one movable object card each."),
    (32, "My School", "Module 4 – Explore My World", "Recognize familiar places and people at school.", "Observe, Environmental Awareness", "Point and talk", "School entrance, classroom, playground and teacher shown in four balanced panels."),
    (33, "My Classroom", "Module 4 – Explore My World", "Find and name common classroom objects.", "Observe, Vocabulary", "Find and circle", "Clean classroom scene containing chair, table, book, pencil, bag, board and blocks."),
    (34, "At the Park", "Module 4 – Explore My World", "Observe playground equipment and safe play actions.", "Observe, Safety Awareness", "Find and tick", "Children using slide, swing and seesaw safely, with one bench, tree and ball."),
    (35, "Community Helpers", "Module 4 – Explore My World", "Match community helpers with their tools.", "Classify, Social Awareness", "Match helper to tool", "Doctor, firefighter, teacher and police officer with stethoscope, hose, book and traffic baton."),
    (36, "Vehicles Around Me", "Module 4 – Explore My World", "Sort transport by land, water and air.", "Classify, Logical Thinking", "Sort vehicles", "Car, bus, bicycle, boat, ship, aeroplane and helicopter with three large category zones."),
    (37, "Inside and Outside", "Module 4 – Explore My World", "Classify familiar objects by where they are usually found.", "Classify, Observe", "Sort inside or outside", "Bed, sofa and lamp mixed with tree, swing and garden ball; two large labelled zones."),
    (38, "Explore My Neighbourhood", "Module 4 – Explore My World", "Follow a simple route and recognize neighbourhood landmarks.", "Observe, Spatial Awareness", "Trace the path", "Very simple map with home, school, park and shop linked by one wide dotted route."),
    (39, "Heavy and Light", "Module 5 – Little Scientists", "Compare familiar objects by likely weight.", "Compare, Predict", "Circle the heavier object", "Four object pairs with obvious weight contrast, such as feather-rock and balloon-watermelon."),
    (40, "Soft and Hard", "Module 5 – Little Scientists", "Classify familiar objects by texture.", "Classify, Sensory Language", "Sort soft or hard", "Pillow, cotton, teddy, stone, wooden block and spoon with two large category areas."),
    (41, "Sink or Float", "Module 5 – Little Scientists", "Make simple predictions about objects in water.", "Predict, Explore", "Predict and tick", "Clear water tub with leaf, stone, cork, spoon and toy boat shown beside two answer columns."),
    (42, "Little Explorer Challenge", "Module 5 – Little Scientists", "Review comparison, classification and prediction.", "Observe, Compare, Predict", "Mixed science challenge", "Four mini-activities reviewing heavy-light, soft-hard and sink-float with large objects."),
    (43, "Explorer Certificate", "Closing", "Celebrate completion and reinforce explorer identity.", "Confidence, Reflection", "Complete certificate", "Premium certificate with approved Star, magnifying-glass emblem, child-name line, date and teacher signature line."),
    (44, "Keep Exploring!", "Closing", "Encourage continued observation, questioning and safe exploration.", "Curiosity, Confidence", "Celebrate and talk", "Back cover with Star and children observing nature; four pillar badges only: Creativity, Communication, Curiosity and Confidence, without descriptions."),
]

NEGATIVE = [
    "no photography or photorealism",
    "no dark full-page background",
    "no unapproved logo, watermark or publisher mark",
    "no empty speech bubbles or invented slogans",
    "no tiny objects, overcrowding or decorative clutter",
    "no distorted faces, hands, limbs or mascot anatomy",
    "no activity content outside the print-safe area",
    "no automatic reconstruction of approved artwork",
]


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.rstrip() + "\n", encoding="utf-8")


def prompt_text(page: tuple) -> str:
    n, title, module, objective, skills, activity, illustration = page
    pid = f"CE-NURSERY-V5-P{n:03d}"
    printed = "hidden" if n <= 5 or n == 44 else str(n - 1)
    return dedent(f"""
    # {pid} — {title}

    ## 1. Release metadata

    - Prompt ID: `{pid}`
    - Version: `5.0.0`
    - Book: Curiosity Explorers Gold
    - Level: Nursery (3+)
    - Physical page: {n} of 44
    - Printed page number: {printed}
    - Module: {module}
    - Exact title: **{title}**
    - Production status: prompt candidate; artwork not approved

    ## 2. Production command

    Create exactly one flat, front-facing A4 portrait workbook page for **{title}**. Follow this package and the locked BCube Gold standards. Produce no collage, alternate page, mock-up, explanation, watermark or unrelated artwork.

    ## 3. Educational metadata

    - Learning objective: {objective}
    - Primary discovery skills: {skills}
    - Activity type: {activity}
    - Developmental domains: cognitive, language, fine-motor and social-emotional as relevant
    - Success evidence: the child completes the visible activity with adult prompting appropriate for age three.

    ## 4. Discovery process

    Use the age-appropriate cycle **Look → Notice → Think → Do → Talk**. Keep the child-facing instruction to one short actionable sentence. Adult guidance belongs in the supporting specification, not as dense page text.

    ## 5. Technical publishing specification

    - Canvas: A4 portrait, 210 × 297 mm
    - Raster target: 2480 × 3508 px at 300 DPI
    - Background: clean white with restrained pastel accents
    - Safe area: all essential text and activity content at least 120 px inside trim
    - Finish: premium preschool publishing, crisp rounded outlines, no heavy shadows
    - Visual/text ratio: approximately 70/30

    ## 6. Layout specification

    - Approved BCube logo at top-left in the locked interior-page position
    - Exact title centred as normal text; one line where practical
    - One rounded instruction banner below the title
    - Main illustration/activity occupies the central page
    - Star appears only when instructionally useful and never covers activity content
    - Standard footer decoration and printed page number policy
    - No Nursery badge on interior pages

    ## 7. Illustration direction

    {illustration}

    Use large recognizable forms, friendly natural expressions, diverse children where people appear, thick clean rounded outlines, restrained depth, and consistent BCube character proportions. Objects required by the activity must be visually unambiguous.

    ## 8. Child activity

    - Interaction: {activity}
    - Child-facing instruction: use one short sentence beginning with an action verb.
    - Provide large touch, circle, matching, tracing or drawing targets suitable for Nursery learners.
    - Do not pre-complete the child's response.

    ## 9. Teacher guidance

    Invite the child to look slowly, name what they notice and explain their choice in a few words. Model positional, descriptive or comparison vocabulary relevant to this page. Accept pointing and one-word responses before expecting a complete sentence.

    ## 10. Parent partnership

    Repeat the core skill with safe household or outdoor objects. Ask one open question, allow observation time and praise the child's effort rather than speed.

    ## 11. Accessibility

    - High figure-ground separation and uncluttered grouping
    - Large readable text and clear icons
    - No colour-only answer dependency where shape or position can reinforce meaning
    - Inclusive, non-stereotyped representation
    - No more than one primary learning task

    ## 12. Design tokens

    Inherit the locked BCube Gold logo, title, instruction-banner, footer, page-number, colour, line-weight and mascot tokens. Do not infer or replace tokens from generated content.

    ## 13. QA validation

    - Exact title, page identity and numbering are correct
    - Required illustration elements and activity targets are present
    - Logo, title, banner and footer follow the approved template
    - Text is grammatically correct and readable at print size
    - No overlaps, clipping, duplicated branding or corrupted elements
    - Output decodes fully at 2480 × 3508 px with 300-DPI metadata
    - Human visual review is recorded before approval

    ## 14. Negative constraints

    {chr(10).join('- ' + item for item in NEGATIVE)}

    ## 15. Acceptance criteria

    Accept only when curriculum, educational, illustration, typography, layout, branding, accessibility and print QA all pass. Mechanical validation alone cannot approve the page. Any failed criterion returns the page to candidate status; never silently reconstruct or substitute content.
    """)


def core_docs() -> dict[str, str]:
    table = "\n".join(f"| P{n:03d} | {title} | {module} | {activity} | {skills} |" for n, title, module, _, skills, activity, _ in PAGES)
    return {
        "00_Project_Overview.md": dedent("""
        # Project Overview

        Curiosity Explorers Gold is Nursery Book 3 in the BCube Future Skills Learning Series. It develops observation, comparison, classification, prediction, safe exploration and questioning through one clear activity per physical page.

        ## Product contract

        - 44 individual physical pages
        - A4 portrait, 2480 × 3508 px, 300 DPI
        - white-background BCube Gold interior system
        - curriculum and prompt identity locked before artwork
        - no page reaches approved assets without human visual QA
        - automation validates and packages; it never makes visual design decisions
        """),
        "01_Curriculum.md": dedent("""
        # Curriculum

        ## Learning progression

        1. Observe and compare familiar objects.
        2. Explore nature through leaves, flowers, trees, birds, insects and weather.
        3. Classify animals by habitat, home, family and food.
        4. Recognize places, helpers, objects and transport in the child's world.
        5. Predict and compare simple physical properties.

        ## Discovery skills

        - Observe: notice visible details
        - Compare: identify similarity and difference
        - Classify: group by one obvious property
        - Explore: investigate safely through guided play
        - Predict: make a simple guess before seeing an outcome
        - Question: use what, where, which and how prompts

        The curriculum is designed for short adult-guided sessions and accepts pointing, naming and simple phrases as valid Nursery responses.
        """),
        "02_Book_Structure.md": dedent("""
        # Book Structure

        - P001–P006: front matter and explorer identity
        - P007–P014: Module 1 — Observe and Compare
        - P015–P022: Module 2 — Nature Explorers
        - P023–P030: Module 3 — Animal Explorers
        - P031–P038: Module 4 — Explore My World
        - P039–P042: Module 5 — Little Scientists
        - P043–P044: certificate and back cover

        P001 and P044 are unnumbered covers. P002–P005 keep printed numbers hidden. Interior numbering follows the locked physical-to-printed mapping in the release manifest.
        """),
        "03_Scope_and_Sequence.md": "# Scope and Sequence\n\n| Page | Exact title | Module | Activity | Discovery skills |\n|---|---|---|---|---|\n" + table,
        "04_Illustration_Bible.md": dedent("""
        # Illustration Bible

        Use a consistent premium 2D preschool illustration system: clean white ground, pastel accents, thick rounded outlines, large recognizable forms, friendly natural expressions and minimal shadows. Keep scenes simple enough for a three-year-old to scan.

        ## Characters

        Children are diverse, age-appropriate and naturally proportioned. Adults are warm and supportive. Avoid exaggerated makeup, fashion poses or adult-like body language.

        ## Objects and nature

        Activity-critical objects must be large, isolated and unambiguous. Plants and animals may be stylized but must retain recognizable defining features. Never use decorative clutter to fill white space.

        ## Continuity

        Reuse approved character and mascot references. Do not reinterpret the BCube logo or Star. Cover artwork must be unique to this book and must not duplicate another series cover.
        """),
        "05_Character_Bible.md": dedent("""
        # Character Bible

        ## Star the Explorer

        Use only the approved Star mascot reference. Lock point count, face, eye spacing, smile, gloves, shoes, cape, colours and outline weight. Appropriate poses include observing, pointing, thinking, encouraging and celebrating. Star must never dominate an activity page or cover required response space.

        ## Explorer children

        Nursery-age children use simple practical clothing, clear gestures and natural facial expressions. Maintain consistent head-to-body proportions and hand anatomy. Show inclusive skin tones, hair textures and abilities without stereotypes.

        ## Adults

        Teachers and parents guide rather than complete activities. Their gaze and gestures should direct attention to the learning target.
        """),
        "06_Design_System.md": dedent("""
        # Design System

        ## Interior master

        - approved BCube logo: fixed top-left placement and one approved version
        - page title: centred normal text, consistent family, weight, colour and optical size
        - instruction banner: one rounded component with locked padding, border and text hierarchy
        - main area: generous white space and one dominant activity
        - footer: locked decorative treatment and page-number placement

        ## Rules

        Do not infer the header from page pixels. Do not recreate logos or fonts with image generation. Covers may use book-specific composition; interiors inherit the common BCube Gold master. Four series pillar badges are Creativity, Communication, Curiosity and Confidence, with no descriptions when used on covers or back covers.
        """),
        "07_QA_Bible.md": dedent("""
        # QA Bible

        A page passes only after eight gates: curriculum, educational, illustration, typography, layout, branding, accessibility and print. Automated checks verify identity, count, dimensions, DPI, decoding, filenames, hashes and required metadata. Human review verifies visual hierarchy, correctness of the activity, character integrity, legibility and preschool usability.

        Status values are `candidate`, `rejected`, `visual-review-pending`, `founder-approved` and `released`. Mechanical success can never promote a page beyond `visual-review-pending`.
        """),
        "08_Publishing_Bible.md": dedent("""
        # Publishing Bible

        Release inputs are immutable approved PNGs plus signed QA records and a release manifest. Individual pages are mandatory; collages and contact sheets cannot enter approved assets. Final packaging verifies 44 sequential physical pages, correct numbering policy, A4 dimensions, 300-DPI metadata, safe margins and full decode.

        Copyright metadata, publisher address, edition and contact information must come from the locked publisher source, never invented by an illustration model.
        """),
        "09_Release_Checklist.md": dedent("""
        # Release Checklist

        - [ ] All 44 prompt files exist and match the manifest
        - [ ] Prompt QA has no missing required section
        - [ ] Cover concept is unique and founder-approved
        - [ ] Approved logo and Star asset hashes are locked
        - [ ] Every page has signed human visual QA
        - [ ] Every PNG decodes at 2480 × 3508 with 300-DPI metadata
        - [ ] Printed page numbers match the release manifest
        - [ ] No candidate, rejected, collage or ZIP asset is inside approved pages
        - [ ] Final PDF is assembled only from approved immutable pages
        - [ ] Release manifest and checksums are archived
        """),
    }


def build_workbook(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Prompts"
    headers = ["Prompt ID", "Physical Page", "Printed Page", "Exact Title", "Module", "Learning Objective", "Discovery Skills", "Activity Type", "Illustration Direction", "Prompt File", "Status"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [24, 14, 14, 30, 34, 48, 30, 28, 70, 34, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(WORKBOOK)


def main() -> None:
    for directory in (DOCS, PROMPTS, QA, PUBLISHING):
        directory.mkdir(parents=True, exist_ok=True)

    for name, content in core_docs().items():
        write(DOCS / name, content)

    records = []
    for page in PAGES:
        n, title, module, objective, skills, activity, illustration = page
        pid = f"CE-NURSERY-V5-P{n:03d}"
        file_name = f"P{n:03d}.md"
        content = prompt_text(page)
        write(PROMPTS / file_name, content)
        records.append({
            "Prompt ID": pid,
            "Physical Page": n,
            "Printed Page": "hidden" if n <= 5 or n == 44 else n - 1,
            "Exact Title": title,
            "Module": module,
            "Learning Objective": objective,
            "Discovery Skills": skills,
            "Activity Type": activity,
            "Illustration Direction": illustration,
            "Prompt File": f"production-prompts/{file_name}",
            "Status": "prompt-candidate",
            "sha256": hashlib.sha256(content.encode("utf-8")).hexdigest(),
        })

    manifest = {
        "schema_version": "1.0.0",
        "book_id": "CURIOSITY_EXPLORERS_NURSERY",
        "book_code": "CE",
        "production_standard": "BCube Gold Production v5",
        "physical_page_count": 44,
        "artwork_generation_allowed": False,
        "phase": 3,
        "pages": records,
    }
    write(PUBLISHING / "phase3-manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False))

    csv_path = ROOT / "Curiosity_Explorers_Gold_Production_Prompts_v5.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=[k for k in records[0] if k != "sha256"])
        writer.writeheader()
        for row in records:
            writer.writerow({k: v for k, v in row.items() if k != "sha256"})

    build_workbook(records)

    qa = {
        "phase": 3,
        "expected_prompts": 44,
        "actual_prompts": len(list(PROMPTS.glob("P*.md"))),
        "required_sections": 15,
        "all_prompt_ids_unique": len({r["Prompt ID"] for r in records}) == 44,
        "all_titles_nonempty": all(r["Exact Title"] for r in records),
        "artwork_generation_blocked": True,
        "status": "pass" if len(records) == 44 else "fail",
    }
    write(QA / "phase3-generation-report.json", json.dumps(qa, indent=2))
    write(QA / "README.md", "# Phase 3 QA\n\nThe generated package is documentation-complete only. Artwork remains blocked until prompt review and founder approval.\n")
    print(f"Generated {len(records)} prompt files and Phase 3 documentation under {ROOT}")


if __name__ == "__main__":
    main()
